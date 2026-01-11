from flask import Flask, render_template, request, redirect, url_for, send_file, session
from openpyxl import load_workbook, Workbook
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime
import math, random, tempfile

app = Flask(__name__)
app.secret_key = "nagyon_titkos_kulcs"
app.config["SESSION_PERMANENT"] = False

# =====================================================
# USERS
# =====================================================
USERS = {
    "Szidi": {"password": generate_password_hash("admin123"), "role": "admin"},
    "Zsuzsi": {"password": generate_password_hash("beo123"), "role": "user"}
}

# =====================================================
# AUTH
# =====================================================
@app.before_request
def force_login():
    if not request.path.startswith("/login") and not request.path.startswith("/static"):
        if "user" not in session:
            return redirect(url_for("login"))

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        u = request.form["username"]
        p = request.form["password"]
        if u in USERS and check_password_hash(USERS[u]["password"], p):
            session.clear()
            session["user"] = u
            return redirect(url_for("index"))
        return render_template("login.html", error="Hibás belépési adatok")
    return render_template("login.html")

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

# =====================================================
# DATA
# =====================================================
workers, shows, schedule = [], [], []
assignment_count, last_days = {}, {}

ROLE_RULES = {
    9: {
        "nézőtér beülős": 2,
        "nézőtér csipog": 2,
        "ruhatár bal": 2,
        "ruhatár jobb": 1,
        "ruhatár erkély": 1,
        "jolly joker": 1
    },
    8: {
        "nézőtér beülős": 2,
        "nézőtér csipog": 2,
        "ruhatár bal": 2,
        "ruhatár jobb": 1,
        "jolly joker": 1
    }
}

MAX_CONSECUTIVE = 3

# =====================================================
# HELPERS
# =====================================================
def normalize_date(v):
    return v.strftime("%Y-%m-%d") if isinstance(v, datetime) else str(v)[:10]

def normalize_list(v):
    if not v or (isinstance(v, float) and math.isnan(v)):
        return []
    return [x.strip() for x in str(v).split(",") if x.strip()]

def import_file(file):
    wb = load_workbook(file, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    return [dict(zip(headers, r)) for r in ws.iter_rows(min_row=2, values_only=True)]

def can_work(name, date):
    days = last_days.get(name, [])
    if len(days) < MAX_CONSECUTIVE:
        return True
    last = sorted(days)[-MAX_CONSECUTIVE:]
    d = datetime.strptime(date, "%Y-%m-%d")
    return not all(
        (d - datetime.strptime(ld, "%Y-%m-%d")).days == i + 1
        for i, ld in enumerate(reversed(last))
    )

def pick_worker(candidates, role, prefer_names=None):
    scored = []
    for w in candidates:
        score = assignment_count[w["név"]]

        # 🔴 FINOM ÉK-HÁTRÁNY BEÜLŐSNÉL
        if role == "nézőtér beülős":
            if w.get("ÉK") == "igen":
                if not prefer_names or w["név"] not in prefer_names:
                    score += 1   # nagyon enyhe hátrány

        scored.append((score, w))

    min_score = min(s for s, _ in scored)
    return random.choice([w for s, w in scored if s == min_score])

# =====================================================
# ROUTES
# =====================================================
@app.route("/")
def index():
    return render_template("import.html")

@app.route("/import/workers", methods=["POST"])
def import_workers():
    global workers
    workers = import_file(request.files["file"])
    assignment_count.clear()
    last_days.clear()
    for w in workers:
        assignment_count[w["név"]] = 0
        last_days[w["név"]] = []
    return redirect("/")

@app.route("/import/shows", methods=["POST"])
def import_shows():
    global shows
    shows = import_file(request.files["file"])
    return redirect("/schedule")

@app.route("/schedule")
def generate_schedule():
    global schedule
    schedule = []

    for show in shows:
        try:
            total = int(float(show.get("létszám", 0)))
        except:
            total = 0

        rules = ROLE_RULES.get(total)
        date = normalize_date(show.get("dátum"))
        title = str(show.get("cím", "")).lower()

        show_block = {
            "cím": show.get("cím"),
            "dátum": show.get("dátum"),
            "szerepek": [],
            "hiba": None
        }

        if not rules:
            show_block["hiba"] = f"Nincs szabály erre a létszámra: {show.get('létszám')}"
            schedule.append(show_block)
            continue

        used, ek_used = set(), False
        assigned = {r: [] for r in rules}

        def eligible(w, role):
            if w["név"] in used:
                return False
            if date in normalize_list(w.get("nem_ér_rá")):
                return False
            if not can_work(w["név"], date):
                return False
            if role == "jolly joker" and w.get("ÉK") == "igen":
                return False
            if w.get("ÉK") == "igen" and ek_used:
                return False
            return True

        # ================= BEÜLŐS =================
        for _ in range(rules["nézőtér beülős"]):
            prefer = [
                w for w in workers
                if title in [s.lower() for s in normalize_list(w.get("nézni_akar"))]
                and eligible(w, "nézőtér beülős")
            ]

            pool = prefer if prefer else [
                w for w in workers if eligible(w, "nézőtér beülős")
            ]

            if not pool:
                break

            prefer_names = {w["név"] for w in prefer}
            w = pick_worker(pool, "nézőtér beülős", prefer_names)

            assigned["nézőtér beülős"].append({
                "név": w["név"],
                "watched": w in prefer
            })

            used.add(w["név"])
            assignment_count[w["név"]] += 1
            last_days[w["név"]].append(date)
            if w.get("ÉK") == "igen":
                ek_used = True

        # ================= TÖBBI SZEREP =================
        for role, needed in rules.items():
            if role == "nézőtér beülős":
                continue

            for _ in range(needed):
                pool = [w for w in workers if eligible(w, role)]
                if not pool:
                    break

                w = pick_worker(pool, role)

                assigned[role].append({
                    "név": w["név"],
                    "watched": False
                })

                used.add(w["név"])
                assignment_count[w["név"]] += 1
                last_days[w["név"]].append(date)
                if w.get("ÉK") == "igen":
                    ek_used = True

        for role in rules:
            show_block["szerepek"].append({
                "szerep": role,
                "kért": rules[role],
                "kiosztott": assigned[role]
            })

        schedule.append(show_block)

    return render_template("schedule.html", schedule=schedule, workers=workers)

# =====================================================
# STATS
# =====================================================
@app.route("/stats")
def stats():
    stats = {}
    for w in workers:
        stats[w["név"]] = {
            "összes": 0,
            "beülős": 0,
            "nézős": 0,
            "ÉK": (w.get("ÉK") == "igen")
        }

    for show in schedule:
        for role in show["szerepek"]:
            for p in role["kiosztott"]:
                stats[p["név"]]["összes"] += 1
                if role["szerep"] == "nézőtér beülős":
                    stats[p["név"]]["beülős"] += 1
                if p.get("watched"):
                    stats[p["név"]]["nézős"] += 1

    return render_template("stats.html", stats=stats)

# =====================================================
# EXPORT
# =====================================================
@app.route("/export/xlsx")
def export_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.append([
        "Előadás","Dátum",
        "Beülős1","Beülős2",
        "Csipog1","Csipog2",
        "Jolly",
        "RB1","RB2",
        "RJ","RE"
    ])

    for s in schedule:
        m = {r["szerep"]: [p["név"] for p in r["kiosztott"]] for r in s["szerepek"]}
        ws.append([
            s["cím"], s["dátum"],
            *(m.get("nézőtér beülős", ["",""])),
            *(m.get("nézőtér csipog", ["",""])),
            *(m.get("jolly joker", [""])),
            *(m.get("ruhatár bal", ["",""])),
            *(m.get("ruhatár jobb", [""])),
            *(m.get("ruhatár erkély", [""]))
        ])

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(tmp.name)
    return send_file(tmp.name, as_attachment=True, download_name="beosztas.xlsx")

# =====================================================
# RUN
# =====================================================
if __name__ == "__main__":
    app.run(debug=True)
